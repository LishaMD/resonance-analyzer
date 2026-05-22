"""
structured_extractor.py — Tabular Data Intelligence Layer
Resonance Analyzer / Coherynce

Sits between app.py extraction and chunker.py.
Intercepts xlsx and csv files and produces semantically rich
key-value representations instead of raw tabular text.

Works on any client's data by detecting structural patterns,
not by looking for specific labels or terminology.
"""

import os
import re
from pathlib import Path


# ── STRUCTURAL PATTERN DETECTION ─────────────────────────────────────────────

def detect_time_dimension(headers: list) -> bool:
    """
    Detects if columns represent a time series.
    Looks for month names, quarter labels, year patterns, or date-like strings.
    Works regardless of client terminology.
    """
    time_patterns = [
        r'\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\b',
        r'\bq[1-4]\b',
        r'\b20\d{2}\b',
        r'\b(month|week|quarter|annual|ytd|mtd)\b',
    ]
    header_str = " ".join(str(h).lower() for h in headers if h)
    return any(re.search(p, header_str) for p in time_patterns)


def detect_comparison_structure(headers: list) -> bool:
    """
    Detects if columns contain actual vs. target/budget/plan pairs.
    Works regardless of what the client calls their columns.
    """
    comparison_signals = [
        'target', 'plan', 'budget', 'forecast', 'goal',
        'actual', 'vs', 'variance', 'delta', 'diff'
    ]
    header_str = " ".join(str(h).lower() for h in headers if h)
    return sum(1 for s in comparison_signals if s in header_str) >= 2


def detect_categorical_distribution(rows: list, col_index: int) -> bool:
    """
    Detects if a column contains repeating categorical values
    (like deal stages, product types, status labels).
    """
    if not rows:
        return False
    values = [str(row[col_index]).strip() for row in rows
              if col_index < len(row) and row[col_index]]
    if len(values) < 3:
        return False
    unique_ratio = len(set(values)) / len(values)
    return unique_ratio < 0.5  # More than 50% repeating = categorical


def detect_entity_attributes(headers: list, rows: list) -> bool:
    """
    Detects relational structure — multiple text columns describing entities.
    Typical of CRM exports, contact lists, pipeline reports.
    """
    if not headers or not rows:
        return False
    text_col_count = sum(
        1 for h in headers
        if h and not any(c.isdigit() for c in str(h))
    )
    return text_col_count >= 3


def is_numeric(value) -> bool:
    """Check if a value is numeric (int, float, or numeric string)."""
    if value is None:
        return False
    try:
        float(str(value).replace(',', '').replace('%', '').replace('$', ''))
        return True
    except ValueError:
        return False


def format_value(value) -> str:
    """Format a cell value cleanly for output."""
    if value is None:
        return ""
    s = str(value).strip()
    if not s or s.lower() in ('none', 'null', 'n/a', '-'):
        return ""
    return s


# ── EXTRACTION STRATEGIES ─────────────────────────────────────────────────────

def extract_trend_analysis(headers: list, rows: list, tab_name: str) -> str:
    """
    For time-series data: compute deltas, identify trends, flag anomalies.
    Works on any labeled time series regardless of terminology.
    """
    lines = [f"[{tab_name}] Time-series data:"]

    for row in rows:
        if not row or not row[0]:
            continue
        label = format_value(row[0])
        if not label:
            continue

        numeric_vals = []
        for i, cell in enumerate(row[1:], 1):
            if is_numeric(cell):
                numeric_vals.append((
                    format_value(headers[i]) if i < len(headers) else f"Period {i}",
                    float(str(cell).replace(',', '').replace('%', '').replace('$', ''))
                ))

        if len(numeric_vals) >= 2:
            first_period, first_val = numeric_vals[0]
            last_period, last_val = numeric_vals[-1]

            if first_val != 0:
                change_pct = ((last_val - first_val) / abs(first_val)) * 100
                direction = "increased" if change_pct > 0 else "decreased"
                lines.append(
                    f"{label}: {direction} {abs(change_pct):.1f}% "
                    f"from {first_period} ({first_val:,.0f}) "
                    f"to {last_period} ({last_val:,.0f})"
                )

                # Flag significant anomalies
                if abs(change_pct) > 50:
                    lines.append(
                        f"  ⚠ SIGNIFICANT CHANGE: {label} changed by "
                        f"{change_pct:+.1f}% — potential anomaly requiring investigation"
                    )

                # Detect acceleration in last period
                if len(numeric_vals) >= 3:
                    mid_val = numeric_vals[len(numeric_vals)//2][1]
                    if mid_val != 0:
                        first_half = ((mid_val - first_val) / abs(first_val)) * 100
                        second_half = ((last_val - mid_val) / abs(mid_val)) * 100
                        if abs(second_half) > abs(first_half) * 1.5:
                            lines.append(
                                f"  ⚠ ACCELERATING TREND: {label} trend "
                                f"accelerating in recent periods"
                            )
        elif len(numeric_vals) == 1:
            period, val = numeric_vals[0]
            lines.append(f"{label} ({period}): {val:,.0f}")

    return "\n".join(lines) if len(lines) > 1 else ""


def extract_comparison_analysis(headers: list, rows: list, tab_name: str) -> str:
    """
    For actual vs. target structures: compute variances, flag misses.
    Works on any paired column structure regardless of terminology.
    """
    lines = [f"[{tab_name}] Performance vs. targets:"]

    # Find pairs of columns that look like actual/target
    actual_indices = []
    target_indices = []
    for i, h in enumerate(headers):
        h_lower = str(h).lower() if h else ''
        if any(t in h_lower for t in ['actual', 'current', 'real', 'achieved']):
            actual_indices.append(i)
        elif any(t in h_lower for t in ['target', 'plan', 'budget', 'goal', 'forecast']):
            target_indices.append(i)

    # If no explicit pairs found, try columns 1 and 2
    if not actual_indices and not target_indices and len(headers) >= 3:
        actual_indices = [1]
        target_indices = [2]

    for row in rows:
        if not row or not row[0]:
            continue
        label = format_value(row[0])
        if not label:
            continue

        for act_i, tgt_i in zip(actual_indices, target_indices):
            if act_i < len(row) and tgt_i < len(row):
                actual = row[act_i]
                target = row[tgt_i]
                if is_numeric(actual) and is_numeric(target):
                    act_val = float(str(actual).replace(',', '').replace('%', '').replace('$', ''))
                    tgt_val = float(str(target).replace(',', '').replace('%', '').replace('$', ''))
                    if tgt_val != 0:
                        variance_pct = ((act_val - tgt_val) / abs(tgt_val)) * 100
                        status = "above" if variance_pct > 0 else "below"
                        lines.append(
                            f"{label}: {act_val:,.0f} vs target {tgt_val:,.0f} "
                            f"— {abs(variance_pct):.1f}% {status} target"
                        )
                        if abs(variance_pct) > 15:
                            lines.append(
                                f"  ⚠ SIGNIFICANT MISS: {label} is "
                                f"{variance_pct:+.1f}% vs target"
                            )

    return "\n".join(lines) if len(lines) > 1 else ""


def extract_categorical_summary(headers: list, rows: list,
                                 tab_name: str) -> str:
    """
    For categorical data: count distributions, identify concentrations.
    Works on CRM pipeline, deal stages, product mix, segment data.
    """
    lines = [f"[{tab_name}] Categorical distribution:"]

    # Find categorical columns
    for col_i, header in enumerate(headers):
        if not header:
            continue
        col_values = [
            format_value(row[col_i])
            for row in rows
            if col_i < len(row) and format_value(row[col_i])
        ]
        if not col_values:
            continue

        if detect_categorical_distribution(rows, col_i):
            from collections import Counter
            counts = Counter(col_values)
            total = len(col_values)
            lines.append(f"\n{header} breakdown ({total} records):")
            for category, count in counts.most_common():
                pct = (count / total) * 100
                lines.append(f"  {category}: {count} ({pct:.1f}%)")

                # Flag dominant categories
                if pct > 60:
                    lines.append(
                        f"  ⚠ CONCENTRATION: {pct:.1f}% of records "
                        f"in single category '{category}'"
                    )

        # Look for numeric columns paired with categorical
        elif col_i > 0 and is_numeric(rows[0][col_i] if rows and col_i < len(rows[0]) else None):
            numeric_vals = [
                float(str(row[col_i]).replace(',', '').replace('%', '').replace('$', ''))
                for row in rows
                if col_i < len(row) and is_numeric(row[col_i])
            ]
            if numeric_vals:
                avg = sum(numeric_vals) / len(numeric_vals)
                lines.append(
                    f"{header}: avg {avg:,.1f} across {len(numeric_vals)} records"
                )

    return "\n".join(lines) if len(lines) > 1 else ""


def extract_entity_summary(headers: list, rows: list, tab_name: str) -> str:
    """
    For entity-attribute data (CRM records, contact lists, pipeline reports):
    summarize key patterns across records rather than listing every row.
    """
    lines = [f"[{tab_name}] Record summary ({len(rows)} records):"]

    if not headers or not rows:
        return ""

    # Summarize each column
    for col_i, header in enumerate(headers):
        if not header:
            continue
        col_values = [
            format_value(row[col_i])
            for row in rows
            if col_i < len(row) and format_value(row[col_i])
        ]
        if not col_values:
            lines.append(f"{header}: no data")
            continue

        # Numeric column — show stats
        numeric_vals = []
        for v in col_values:
            try:
                numeric_vals.append(
                    float(v.replace(',', '').replace('%', '').replace('$', ''))
                )
            except ValueError:
                pass

        if len(numeric_vals) > len(col_values) * 0.7:
            avg = sum(numeric_vals) / len(numeric_vals)
            min_v = min(numeric_vals)
            max_v = max(numeric_vals)
            lines.append(
                f"{header}: avg {avg:,.1f} | range {min_v:,.1f}–{max_v:,.1f}"
            )

        # Categorical column — show distribution
        elif detect_categorical_distribution(rows, col_i):
            from collections import Counter
            counts = Counter(col_values)
            top = counts.most_common(3)
            summary = ", ".join(f"{k} ({v})" for k, v in top)
            lines.append(f"{header}: {summary}")

        # Text with missing values — flag gaps
        else:
            missing = len(rows) - len(col_values)
            if missing > 0:
                pct_missing = (missing / len(rows)) * 100
                lines.append(
                    f"{header}: {len(col_values)} populated, "
                    f"{missing} empty ({pct_missing:.0f}% missing)"
                )
                if pct_missing > 30:
                    lines.append(
                        f"  ⚠ DATA QUALITY: {pct_missing:.0f}% of "
                        f"'{header}' records have no value"
                    )

    return "\n".join(lines) if len(lines) > 1 else ""


# ── MAIN EXTRACTION ROUTER ────────────────────────────────────────────────────

def extract_tab(tab_name: str, headers: list, rows: list) -> str:
    """
    Routes a single spreadsheet tab to the appropriate extraction strategy
    based on detected structural patterns.
    Falls back to uniform key-value if no pattern detected.
    """
    if not headers or not rows:
        return ""

    results = []

    # Apply all relevant strategies — a tab can match multiple patterns
    if detect_time_dimension(headers):
        result = extract_trend_analysis(headers, rows, tab_name)
        if result:
            results.append(result)

    if detect_comparison_structure(headers):
        result = extract_comparison_analysis(headers, rows, tab_name)
        if result:
            results.append(result)

    if detect_entity_attributes(headers, rows):
        result = extract_entity_summary(headers, rows, tab_name)
        if result:
            results.append(result)

    # Always run categorical for any tab with enough rows
    if len(rows) >= 5:
        result = extract_categorical_summary(headers, rows, tab_name)
        if result and len(result.split('\n')) > 2:
            results.append(result)

    # Fallback: uniform key-value extraction
    if not results:
        lines = [f"[{tab_name}] Key values:"]
        for row in rows:
            if not row or not row[0]:
                continue
            label = format_value(row[0])
            values = [format_value(cell) for cell in row[1:] if format_value(cell)]
            if label and values:
                lines.append(f"{label}: {' | '.join(values)}")
        if len(lines) > 1:
            results.append("\n".join(lines))

    return "\n\n".join(results)


def extract_structured(file_path: str, filename: str) -> list[dict]:
    """
    Main entry point. Reads an xlsx or csv file and returns a list of
    structured chunk dicts ready for chunker.py and embedder.py.

    Each chunk represents one tab (xlsx) or one record group (csv)
    with semantically rich interpreted content.
    """
    ext = Path(filename).suffix.lower()
    chunks = []

    if ext in ('.xlsx', '.xls'):
        chunks = _extract_xlsx(file_path, filename)
    elif ext == '.csv':
        chunks = _extract_csv(file_path, filename)

    return chunks


def _extract_xlsx(file_path: str, filename: str) -> list[dict]:
    """Extract and interpret each tab of an xlsx file."""
    from openpyxl import load_workbook
    chunks = []

    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"  [structured_extractor] Failed to open {filename}: {e}")
        return []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))

        if not all_rows:
            continue

        # First non-empty row is headers
        headers = []
        data_rows = []
        found_headers = False

        for row in all_rows:
            if not found_headers:
                if any(cell is not None for cell in row):
                    headers = [str(cell).strip() if cell is not None else ""
                               for cell in row]
                    found_headers = True
            else:
                if any(cell is not None for cell in row):
                    data_rows.append(list(row))

        if not headers or not data_rows:
            continue

        interpreted = extract_tab(sheet_name, headers, data_rows)

        if interpreted:
            chunks.append({
                "filename": filename,
                "extracted_text": interpreted,
                "tab_name": sheet_name,
                "doc_type_override": "spreadsheet"
            })
            print(f"  [structured_extractor] {filename} / {sheet_name} → interpreted")

    return chunks


def _extract_csv(file_path: str, filename: str) -> list[dict]:
    """Extract and interpret a CSV file."""
    import csv
    chunks = []

    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.reader(f)
            all_rows = list(reader)
    except Exception as e:
        print(f"  [structured_extractor] Failed to open {filename}: {e}")
        return []

    if not all_rows:
        return []

    headers = [str(h).strip() for h in all_rows[0]]
    data_rows = [row for row in all_rows[1:] if any(cell.strip() for cell in row)]

    if not data_rows:
        return []

    # Split large CSVs into groups of 50 rows for better retrieval
    group_size = 50
    for i in range(0, len(data_rows), group_size):
        group = data_rows[i:i + group_size]
        tab_label = f"records {i+1}–{min(i+group_size, len(data_rows))}"
        interpreted = extract_tab(tab_label, headers, group)

        if interpreted:
            chunks.append({
                "filename": filename,
                "extracted_text": interpreted,
                "tab_name": tab_label,
                "doc_type_override": "crm"
            })
            print(f"  [structured_extractor] {filename} / {tab_label} → interpreted")

    return chunks


# ── STANDALONE TEST ───────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys

    # Test with TerraLoop files if available
    test_files = [
        "/Users/elishadavison/Desktop/resonance-analyzer/TerraLoop Documents/TerraLoop FinancialModel.xlsx",
        "/Users/elishadavison/Desktop/resonance-analyzer/TerraLoop Documents/TerraLoop Pipeline CRM Export.csv",
    ]

    for file_path in test_files:
        if not Path(file_path).exists():
            print(f"File not found: {file_path}")
            continue

        filename = Path(file_path).name
        print(f"\n{'='*60}")
        print(f"Testing: {filename}")
        print('='*60)

        chunks = extract_structured(file_path, filename)
        print(f"\nProduced {len(chunks)} interpreted chunks:\n")

        for i, chunk in enumerate(chunks):
            print(f"── Chunk {i+1}: {chunk['tab_name']} ──")
            print(chunk['extracted_text'][:500])
            print("..." if len(chunk['extracted_text']) > 500 else "")
            print()